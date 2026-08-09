#!/usr/bin/env python3
"""Memory-safe compiler for the supplied iTRAQ telemetry workbook.

GitHub Pages cannot read a workstation Excel file at runtime. This script streams
the workbook twice and publishes a compact JavaScript payload containing only
facts that can be calculated from its telemetry columns. It never creates driver,
HR, customer, order, fatigue, seatbelt, or punctuality data because those fields
do not exist in the source workbook.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(
    "/Users/chenyixuan/Dev/比賽/hino/"
    "HINO x GenAI：運用商用車車聯網大數據，發展商用車頭家的智慧夥伴/"
    "output data_Hotai_20260511.xlsx"
)
DEFAULT_OUTPUT = ROOT / "excel-derived-data.js"
REGION_INFO = {
    "N": ("北區", "北部", "#38bdf8"),
    "C": ("中區", "中部", "#fb923c"),
    "S": ("南區", "南部", "#34d399"),
}
METRIC_INFO = {
    "safety": ("計算安全分", "分", True, "分數越高代表超速、怠速、高引擎負載與 DTC 記錄較低"),
    "speed": ("超速紀錄", "筆", False, "GPS 車速高於該筆記錄限速；越少越好"),
    "idle": ("怠速佔比", "%", False, "carStatus=2 /（行駛 + 怠速）記錄；越低越好"),
    "load": ("高引擎負載紀錄", "筆", False, "CAN engineLoad ≥ 90 的記錄；越少越好"),
    "dtc": ("DTC 紀錄", "筆", False, "事件欄位帶有 DTC code 的記錄；越少越好"),
    "fuel": ("百公里油耗", "L", False, "以每車 CAN 累積油耗與里程差分計算；越低越好"),
}


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float | None:
    try:
        return float(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def whole(value: Any, default: int = -1) -> int:
    value = number(value)
    return int(value) if value is not None else default


def stamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def pretty(value: float | None, digits: int = 1) -> float | int:
    if value is None:
        return 0
    result = round(value, digits)
    return int(result) if digits == 0 else result


def ratio(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


@dataclass
class Stats:
    rows: int = 0
    driving: int = 0
    idling: int = 0
    overspeed: int = 0
    high_load: int = 0
    dtc: int = 0
    counters: dict[str, dict[str, tuple[datetime, float]]] = field(default_factory=dict)

    def add(self, car: str, when: datetime, status: int, speed: float | None, limit: float | None,
            load: float | None, dtc: bool, fuel: float | None, mileage: float | None) -> None:
        self.rows += 1
        self.driving += status == 1
        self.idling += status == 2
        self.overspeed += bool(limit and limit > 0 and speed is not None and speed > limit)
        self.high_load += bool(load is not None and load >= 90)
        self.dtc += bool(dtc)
        car_counts = self.counters.setdefault(car, {})
        for key, value in (("fuel", fuel), ("mileage", mileage)):
            if value is None:
                continue
            old = car_counts.get(key)
            if old is None:
                car_counts[key] = ((when, value), (when, value))
            else:
                first, last = old
                car_counts[key] = ((when, value) if when < first[0] else first,
                                   (when, value) if when > last[0] else last)

    def idle_pct(self) -> float:
        return pretty(ratio(self.idling * 100, self.driving + self.idling), 1)

    def overspeed_pct(self) -> float:
        return pretty(ratio(self.overspeed * 100, self.driving), 1)

    def high_load_pct(self) -> float:
        return pretty(ratio(self.high_load * 100, self.driving), 1)

    def score(self) -> int:
        # Clearly marked in the UI as a calculated prototype index—not HINO's official score.
        penalty = min(42, self.overspeed_pct() * 1.15)
        penalty += min(24, self.idle_pct() * 0.7)
        penalty += min(16, self.high_load_pct() * 1.1)
        penalty += min(10, ratio(self.dtc * 1000, self.rows) * 0.25)
        return max(0, min(100, round(100 - penalty)))

    def fuel_per_100km(self) -> float:
        fuel_delta = 0.0
        km_delta = 0.0
        for values in self.counters.values():
            if "fuel" not in values or "mileage" not in values:
                continue
            fuel_delta += max(0.0, values["fuel"][1][1] - values["fuel"][0][1])
            km_delta += max(0.0, values["mileage"][1][1] - values["mileage"][0][1])
        return pretty(ratio(fuel_delta * 100, km_delta), 2)


def workbook_rows(path: Path):
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book["output"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    positions = {str(name): index for index, name in enumerate(headers)}
    required = ["journeyCode", "time", "carStatus", "carNum", "gps.longitude", "gps.latitude",
                "gps.speed", "gps.speedLimit", "can.totalMileage", "can.engine.totalFuelUsed",
                "can.engine.engineLoad", "event[0].info.dtcCodes[0]",
                "event[1].info.dtcCodes[0]", "event[2].info.dtcCodes[0]"]
    for raw in rows:
        yield {key: raw[positions[key]] if key in positions else None for key in required}
    book.close()


def region_id(latitude: float | None) -> str:
    if latitude is None:
        return "C"
    return "N" if latitude >= 25 else "C" if latitude >= 24 else "S"


def month_values(stats: list[Stats], metric: str) -> list[float | int]:
    getter = {
        "safety": lambda item: item.score(),
        "speed": lambda item: item.overspeed,
        "idle": lambda item: item.idle_pct(),
        "load": lambda item: item.high_load,
        "dtc": lambda item: item.dtc,
        "fuel": lambda item: item.fuel_per_100km(),
    }[metric]
    return [getter(item) for item in stats]


def issue(vehicle: dict) -> str:
    ranked = sorted([
        (vehicle["overspeed_pct"], f"超速 {vehicle['overspeed_count']:,} 筆"),
        (vehicle["idle_pct"], f"怠速 {vehicle['idle_pct']}%"),
        (vehicle["high_load_pct"], f"高引擎負載 {vehicle['high_load_count']:,} 筆"),
        (vehicle["dtc_count"], f"DTC {vehicle['dtc_count']:,} 筆"),
    ], reverse=True)
    return "／".join(value for score, value in ranked if score > 0)[:80] or "無可判讀事件"


def metric_payload(global_months: list[Stats]) -> tuple[list[dict], dict, dict]:
    metrics: list[dict] = []
    facts: dict[str, str] = {}
    solutions: dict[str, dict] = {}
    actions = {
        "safety": "檢視高風險車號並安排安全提醒",
        "speed": "對超速紀錄高的車號建立限速提醒",
        "idle": "對怠速偏高車號建立熄火提醒",
        "load": "安排高引擎負載車輛檢修與派車覆核",
        "dtc": "建立 DTC 檢修優先清單",
        "fuel": "針對高油耗車檢查怠速、路線與保養",
    }
    for key, (name, unit, high, hint) in METRIC_INFO.items():
        data = month_values(global_months, key)
        facts[key] = f"2025 年 11 月全隊{name} {data[-1]:,}{unit}。"
        if key == "safety":
            facts[key] = f"2025 年 11 月全隊計算安全分 {data[-1]} 分；計算只使用 Excel 的超速、怠速、高引擎負載與 DTC 欄位。"
        if key == "idle":
            facts[key] = f"2025 年 11 月全隊怠速佔比 {data[-1]}%，以 carStatus=2 計算。"
        if key == "fuel":
            facts[key] = f"2025 年 11 月百公里油耗 {data[-1]} L，以每車 CAN 累積油耗與里程差分計算。"
        metrics.append({
            "key": key, "name": name, "unit": unit, "data": data,
            "cause": {"t": f"{name}為何需要關注？", "rs": [
                "資料只反映車聯網紀錄；需要再核對路況、車況與派車條件。",
                "先以車號下鑽，再決定是否需要提醒、保修或調度。",
            ], "act": [{"l": actions[key], "c": "pri", "fn": "act('已建立 Excel 車聯網資料的後續處理清單。','ok')"}]},
        })
        solutions[key] = {"t": f"{name}偏高／偏低，怎麼處理？", "pts": [
            "先確認原始時間、車號、路況與車況，避免以單一訊號歸因於駕駛。",
            actions[key] + "，處置結果需回填後再評估。",
            "不得以單一計算分數自動扣薪、裁員或解約。",
        ], "btn": [{"l": "建立處理清單", "c": "pri", "fn": "act('已建立資料覆核與處理清單。','ok')"}]}
    return metrics, facts, solutions


def build(source: Path) -> dict:
    latest: dict[str, dict] = {}
    car_regions: dict[str, str] = {}
    earliest: datetime | None = None
    newest: datetime | None = None
    records = 0
    # First streamed pass: vehicle ownership area comes only from latest GPS latitude.
    for row in workbook_rows(source):
        car, when = text(row["carNum"]), stamp(row["time"])
        if not car or not when:
            continue
        records += 1
        earliest = when if earliest is None or when < earliest else earliest
        newest = when if newest is None or when > newest else newest
        if car not in latest or when > latest[car]["time"]:
            latest[car] = {"time": when, "latitude": number(row["gps.latitude"]), "longitude": number(row["gps.longitude"]),
                           "speed": number(row["gps.speed"]), "limit": number(row["gps.speedLimit"]), "status": whole(row["carStatus"]),
                           "journey": text(row["journeyCode"])}
    for car, row in latest.items():
        car_regions[car] = region_id(row["latitude"])

    all_months = [Stats() for _ in range(11)]
    region_all = {key: Stats() for key in REGION_INFO}
    region_months = {key: [Stats() for _ in range(11)] for key in REGION_INFO}
    vehicle_stats: dict[str, Stats] = defaultdict(Stats)
    car_journeys: dict[str, set[str]] = defaultdict(set)
    total_journeys: set[str] = set()
    # Second pass: metrics and counter deltas.
    for row in workbook_rows(source):
        car, when = text(row["carNum"]), stamp(row["time"])
        if not car or not when or not 1 <= when.month <= 11:
            continue
        status, speed, limit, load = whole(row["carStatus"]), number(row["gps.speed"]), number(row["gps.speedLimit"]), number(row["can.engine.engineLoad"])
        dtc = any(text(row[key]) for key in ("event[0].info.dtcCodes[0]", "event[1].info.dtcCodes[0]", "event[2].info.dtcCodes[0]"))
        fuel, mileage = number(row["can.engine.totalFuelUsed"]), number(row["can.totalMileage"])
        region, month = car_regions[car], when.month - 1
        for stat in (all_months[month], region_all[region], region_months[region][month], vehicle_stats[car]):
            stat.add(car, when, status, speed, limit, load, dtc, fuel, mileage)
        journey = text(row["journeyCode"])
        if journey:
            car_journeys[car].add(journey)
            total_journeys.add(journey)

    vehicles: list[dict] = []
    for car, stat in vehicle_stats.items():
        last = latest[car]
        vehicle = {
            "n": car, "c": car, "s": stat.score(), "region": car_regions[car],
            "overspeed_count": stat.overspeed, "overspeed_pct": stat.overspeed_pct(),
            "idle_count": stat.idling, "idle_pct": stat.idle_pct(),
            "high_load_count": stat.high_load, "high_load_pct": stat.high_load_pct(), "dtc_count": stat.dtc,
            "last_time": last["time"].strftime("%Y-%m-%d %H:%M:%S"),
            "last_status": {0: "停車", 1: "行駛中", 2: "怠速"}.get(last["status"], "未提供"),
            "last_speed": pretty(last["speed"], 0), "last_limit": pretty(last["limit"], 0),
            "position": f"{last['longitude']:.6f}, {last['latitude']:.6f}" if last["longitude"] is not None and last["latitude"] is not None else "未提供",
            "journeys": len(car_journeys[car]), "journey": last["journey"],
        }
        vehicle["i"] = issue(vehicle)
        vehicles.append(vehicle)
    regions: list[dict] = []
    code_by_car: dict[str, str] = {}
    for region, (name, scope, color) in REGION_INFO.items():
        drivers = sorted((item for item in vehicles if item["region"] == region), key=lambda item: (item["s"], item["c"]))
        if not drivers:
            continue
        for index, vehicle in enumerate(drivers):
            code_by_car[vehicle["c"]] = f"{region}{index}"
        stat = region_all[region]
        series = {key: month_values(region_months[region], key) for key in METRIC_INFO}
        regions.append({
            "id": region, "name": name, "scope": scope, "lead": "原始資料未提供", "phone": "原始資料未提供", "color": color,
            "drivers": drivers, "idlePct": stat.idle_pct(), "overload": stat.high_load, "seatbelt": stat.dtc,
            "fuel": stat.fuel_per_100km(), "anomaly": pretty(ratio(stat.overspeed + stat.high_load + stat.dtc, stat.rows) * 100, 1),
            "onTime": None, "journeys": sum(item["journeys"] for item in drivers),
            "safety": series["safety"], "brake": series["speed"], "series": series,
        })
    metrics, facts, solutions = metric_payload(all_months)
    total_series = {key: month_values(all_months, key) for key in METRIC_INFO}
    by_speed = sorted(vehicles, key=lambda item: item["overspeed_count"], reverse=True)[:3]
    by_idle = sorted(vehicles, key=lambda item: item["idle_pct"], reverse=True)[:3]
    by_load = sorted(vehicles, key=lambda item: item["high_load_count"], reverse=True)[:3]
    by_dtc = sorted(vehicles, key=lambda item: item["dtc_count"], reverse=True)[:3]
    labels = lambda items, key: "、".join(f"{item['c']} {item[key]:,} 筆" for item in items)
    todo_specs = [
        ("bad", "超速風險", by_speed, "overspeed_count", "超速紀錄最高", "GPS 車速高於該筆限速"),
        ("warn", "怠速改善", by_idle, "idle_count", "怠速比例偏高", "carStatus=2"),
        ("warn", "高引擎負載", by_load, "high_load_count", "高引擎負載紀錄較多", "CAN engineLoad ≥ 90"),
        ("info", "DTC 檢修", by_dtc, "dtc_count", "有 DTC 紀錄", "event DTC code"),
    ]
    todos, todo_data = [], []
    for severity, category, items, key, headline, source_label in todo_specs:
        relevant = [item for item in items if item[key] > 0]
        codes = [code_by_car[item["c"]] for item in relevant]
        todos.append({"sev": severity, "cat": category, "decide": True,
                      "tt": f"{len(relevant)} 台車{headline}，請先做資料覆核",
                      "dt": labels(relevant, key) if relevant else "資料期間未發現此類紀錄",
                      "status": {"by": "AI", "txt": f"已依 Excel 的 {source_label} 欄位建立待覆核清單。"},
                      "acts": [{"l": "查看車號", "c": "gho", "fn": f"drillDrivers({json.dumps(codes, ensure_ascii=False)},'Excel {category} 車號')"}]})
        todo_data.append(f"<b>資料依據：</b>{source_label}。<br><b>下一步：</b>先核對車況、路況與派車情境；不可據此直接做薪酬、裁員或解約決定。")
    advice = [
        {"f": f"<b>{by_speed[0]['c']}</b> 的超速紀錄最高", "why": f"Excel 期間共 {by_speed[0]['overspeed_count']:,} 筆，佔行駛紀錄 {by_speed[0]['overspeed_pct']}%。", "status": {"by": "AI", "txt": "可建立限速提醒草稿；請先確認限速資料與路況。"}, "acts": [{"l": "建立提醒草稿", "c": "pri", "fn": "act('已建立以車號為對象的限速提醒草稿。','ok')"}]},
        {"f": f"<b>{by_idle[0]['c']}</b> 的怠速比例最高", "why": f"Excel 期間怠速 {by_idle[0]['idle_pct']}%，共 {by_idle[0]['idle_count']:,} 筆 carStatus=2 紀錄。", "status": {"by": "AI", "txt": "可建立熄火提醒；來源未提供裝卸、等待或排班原因。"}, "acts": [{"l": "建立熄火提醒", "c": "pri", "fn": "act('已建立以車號為對象的怠速提醒草稿。','ok')"}]},
        {"f": f"<b>{by_load[0]['c']}</b> 的高引擎負載紀錄最高", "why": f"CAN engineLoad ≥ 90 共 {by_load[0]['high_load_count']:,} 筆，需搭配車況與派車資料判讀。", "status": {"by": "AI", "txt": "已列為車況覆核候選，不將此訊號直接等同超載或人員績效。"}, "acts": [{"l": "建立覆核", "c": "pri", "fn": "act('已建立車況與派車覆核任務。','ok')"}]},
    ]
    latest_vehicles = sorted(vehicles, key=lambda item: item["last_time"], reverse=True)
    orders = [{"id": item["journey"] or f"車號-{item['c']}", "car": item["c"], "status": item["last_status"], "cur": f"最後紀錄 {item['last_time']}",
               "from": "起始 GPS", "to": "最後 GPS", "etaMin": None, "progress": 1, "risk": bool(item["overspeed_count"] or item["idle_count"]), "driver": "原始資料未提供", "last_speed": item["last_speed"], "shipper": "telemetry"} for item in latest_vehicles[:4]]
    first_region, last_region = regions[0], regions[-1]
    return {
        "meta": {"sourceFile": source.name, "records": records, "vehicles": len(vehicles), "period": f"{earliest:%Y-%m-%d} 至 {newest:%Y-%m-%d}", "lastRecord": f"{newest:%Y-%m-%d %H:%M:%S}", "driverFieldsAvailable": False, "orderFieldsAvailable": False, "scoreMethod": "100 − 超速率、怠速率、高引擎負載率與 DTC 記錄的上限加權扣分"},
        "months": [f"{month}月" for month in range(1, 12)], "regions": regions,
        "aggregate": {"safety": total_series["safety"], "idlePct": total_series["idle"][-1], "fuel": total_series["fuel"][-1], "journeys": len(total_journeys)},
        "ordersByRegion": {item["id"]: item["journeys"] for item in regions}, "targetJourneysPerVehicle": max(1, round(len(total_journeys) / len(vehicles))),
        "metrics": metrics, "factMap": facts, "dims": [{"key": key, "name": info[0], "high": info[2], "unit": info[1], "hint": info[3]} for key, info in METRIC_INFO.items()], "dimSolData": solutions,
        "todos": todos, "todoData": todo_data, "advice": advice, "shippers": [{"id": "telemetry", "name": "Excel 車聯網紀錄", "orders": orders}],
        "accountBindings": {"lead_region": first_region["id"], "driver_code": f"{first_region['id']}0", "personal_code": f"{last_region['id']}0"}, "vehicleSnapshot": latest_vehicles,
        "sourceNote": "本頁數據由 output data_Hotai_20260511.xlsx 計算；原始檔未提供駕駛姓名、工時、人資、訂單、準時率與安全帶欄位。",
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    payload = build(args.source)
    args.output.write_text("/* Generated from the supplied iTRAQ workbook. */\nwindow.HINO_EXCEL_DATA = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(f"Wrote {args.output} ({payload['meta']['records']:,} telemetry rows / {payload['meta']['vehicles']} vehicles)")


if __name__ == "__main__":
    main()
